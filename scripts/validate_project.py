from __future__ import annotations

import argparse
import hashlib
import json
import re
import subprocess
import sys
from pathlib import Path
from zipfile import ZipFile

import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from scripts.check_notebook import inspect_notebook
from src.experiment_matrix import build_experiment_matrix

ARTIFACTS = ROOT / "artifacts"


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1 << 20), b""):
            digest.update(chunk)
    return digest.hexdigest()


def read_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def validate_markdown_links(path: Path) -> list[str]:
    text = path.read_text(encoding="utf-8")
    broken = []
    for target in re.findall(r"\[[^\]]+\]\(([^)]+)\)", text):
        clean = target.split("#", 1)[0]
        if not clean or re.match(r"^[a-z]+://", clean, re.IGNORECASE):
            continue
        if not (path.parent / clean).resolve().exists():
            broken.append(target)
    return broken


def core_checks() -> tuple[dict, list[str]]:
    errors: list[str] = []
    required = [
        "README.md",
        "BTL_XuLyAnh_NhanDienHoa.ipynb",
        "streamlit_app.py",
        ".streamlit/config.toml",
        "configs/experiment.yaml",
        "configs/degradation_matrix.json",
        "src/preprocessing.py",
        "src/degradations.py",
        "src/enhancements.py",
        "src/model.py",
        "src/evaluate.py",
        "app_components/pipeline.py",
        "Dockerfile",
        "docs/STREAMLIT_DEPLOYMENT.md",
        "docs/MODEL_CARD.md",
        "docs/DATA_CARD.md",
        "docs/EXPERIMENTS.md",
        "LICENSE",
        "QUALITY_REPORT.md",
    ]
    missing = [path for path in required if not (ROOT / path).exists()]
    errors.extend(f"Missing core file: {path}" for path in missing)

    split_paths = {name: ROOT / "splits" / f"{name}.csv" for name in ("train", "validation", "test")}
    frames = {name: pd.read_csv(path) for name, path in split_paths.items() if path.exists()}
    split_checks: dict[str, int] = {}
    expected_counts = {"train": 2571, "validation": 549, "test": 550}
    if set(frames) != set(split_paths):
        errors.append("One or more split files are missing")
    else:
        for name, frame in frames.items():
            missing_columns = {"relative_path", "label", "sha256"}.difference(frame.columns)
            if missing_columns:
                errors.append(f"{name}.csv missing columns: {sorted(missing_columns)}")
            if len(frame) != expected_counts[name]:
                errors.append(f"{name}.csv count is {len(frame)}, expected {expected_counts[name]}")
        for left, right in (("train", "validation"), ("train", "test"), ("validation", "test")):
            path_overlap = len(set(frames[left].relative_path) & set(frames[right].relative_path))
            hash_overlap = len(set(frames[left].sha256) & set(frames[right].sha256))
            split_checks[f"{left}_{right}_path_overlap"] = path_overlap
            split_checks[f"{left}_{right}_hash_overlap"] = hash_overlap
            if path_overlap or hash_overlap:
                errors.append(f"Leakage detected for {left}/{right}: paths={path_overlap}, hashes={hash_overlap}")

    matrix = build_experiment_matrix()
    if len(matrix) != 49 or len({item.condition_id for item in matrix}) != 49:
        errors.append("Experiment matrix is not exactly 49 unique conditions")

    if (ROOT / "backend").exists() or (ROOT / "frontend").exists() or (ROOT / "docker-compose.yml").exists():
        errors.append("Legacy two-service production paths still exist outside legacy/")
    streamlit_source = (ROOT / "streamlit_app.py").read_text(encoding="utf-8") if (ROOT / "streamlit_app.py").exists() else ""
    if "requests" in streamlit_source or "http://localhost" in streamlit_source:
        errors.append("Standalone Streamlit app still contains an internal HTTP dependency")
    if "@st.cache_resource" not in streamlit_source:
        errors.append("Streamlit app does not cache the model resource")

    inventory_path = ROOT / "data" / "inventory.csv"
    inventory_rows = len(pd.read_csv(inventory_path)) if inventory_path.exists() else 0
    if inventory_rows != 3670:
        errors.append(f"Canonical data/inventory.csv has {inventory_rows} rows; expected 3670")
    if (ROOT / "models" / "surrogate_classifier.joblib").exists():
        errors.append("Surrogate classifier is forbidden from the canonical final package")
    if (ROOT / "results" / "dataset_inventory.csv").exists():
        errors.append("Duplicate inventory exists at results/dataset_inventory.csv")

    readme_path = ROOT / "README.md"
    readme_lines = len(readme_path.read_text(encoding="utf-8").splitlines()) if readme_path.exists() else 0
    if not 60 <= readme_lines <= 320:
        errors.append(f"README must contain 60-320 focused lines; found {readme_lines}")
    broken_links = validate_markdown_links(readme_path) if readme_path.exists() else []
    if broken_links:
        errors.append(f"README has broken local links: {broken_links}")

    status = {
        "missing_core_files": missing,
        "split_counts": {name: len(frame) for name, frame in frames.items()},
        "split_checks": split_checks,
        "experiment_condition_count": len(matrix),
        "standalone_streamlit": not any("Legacy two-service" in error for error in errors),
        "canonical_inventory_rows": inventory_rows,
        "readme_lines": readme_lines,
        "readme_broken_local_links": broken_links,
    }
    return status, errors


def office_visible_text(path: Path) -> str:
    with ZipFile(path) as archive:
        return " ".join(
            re.sub(r"<[^>]+>", " ", archive.read(name).decode("utf-8", errors="ignore"))
            for name in archive.namelist()
            if name.endswith(".xml")
        )


def full_run_checks(require_submission: bool = False) -> tuple[dict, list[str]]:
    errors: list[str] = []
    required = [
        "models/mobilenetv2_flowers.keras",
        "configs/locked_enhancement_params.json",
        "artifacts/environment.json",
        "artifacts/training/history.csv",
        "artifacts/training/learning_curves.png",
        "results/final/predictions.csv",
        "results/final/condition_metrics.csv",
        "results/final/per_class_metrics.csv",
        "results/final/statistical_tests.csv",
        "results/final/manifest.json",
    ]
    if require_submission:
        required.extend([
            "artifacts/app_smoke_test.json",
            "artifacts/deployment_verification.json",
            "configs/submission_metadata.json",
            "docs/REPORT_FINAL.docx",
            "docs/REPORT_FINAL.pdf",
            "docs/SLIDES_FINAL.pptx",
            "docs/ASSIGNMENT_FINAL.xlsx",
        ])
    missing = [path for path in required if not (ROOT / path).exists()]
    errors.extend(f"Missing final artifact: {path}" for path in missing)

    condition_path = ROOT / "results" / "final" / "condition_metrics.csv"
    if condition_path.exists():
        condition = pd.read_csv(condition_path)
        expected_ids = {item.condition_id for item in build_experiment_matrix()}
        if len(condition) != 49 or "condition_id" not in condition or condition.get("condition_id", pd.Series(dtype=str)).duplicated().any():
            errors.append("condition_metrics.csv must contain 49 unique condition_id rows")
        elif set(condition["condition_id"]) != expected_ids:
            errors.append("condition_metrics.csv condition IDs differ from the locked matrix")
        required_metrics = {
            "accuracy", "macro_precision", "macro_recall", "macro_f1", "weighted_f1",
            "psnr", "ssim", "delta_e_2000", "inference_time_ms_per_image_mean",
            "inference_time_ms_per_image_median", "inference_time_ms_per_image_p95", "latency_runs",
        }
        if not required_metrics.issubset(condition.columns):
            errors.append(f"condition_metrics.csv missing metrics: {sorted(required_metrics.difference(condition.columns))}")
        else:
            always_finite = required_metrics.difference({"delta_e_2000"})
            if condition[list(always_finite)].isna().any().any():
                errors.append("condition_metrics.csv contains missing required metric values")
            delta_required = condition["degradation"].eq("color_cast") | condition["image_type"].eq("clean")
            if condition.loc[delta_required, "delta_e_2000"].isna().any():
                errors.append("Color-cast/clean conditions require delta_e_2000")
        if "sample_count" not in condition or not condition["sample_count"].eq(550).all():
            errors.append("Every condition_metrics.csv row must document sample_count=550")

    prediction_path = ROOT / "results" / "final" / "predictions.csv"
    if prediction_path.exists():
        predictions = pd.read_csv(prediction_path)
        if len(predictions) != 49 * 550:
            errors.append(f"predictions.csv has {len(predictions)} rows; FULL_RUN requires {49 * 550}")
        required_prediction_columns = {
            "run_id", "condition_id", "relative_path", "sha256", "true_label",
            "predicted_label", "confidence", "correct", "probabilities_json",
        }
        if not required_prediction_columns.issubset(predictions.columns):
            errors.append(f"predictions.csv missing columns: {sorted(required_prediction_columns.difference(predictions.columns))}")
        else:
            if predictions.duplicated(["condition_id", "relative_path"]).any():
                errors.append("predictions.csv contains duplicate condition/path pairs")
            counts = predictions.groupby("condition_id").size()
            if len(counts) != 49 or not counts.eq(550).all():
                errors.append("predictions.csv must contain exactly 550 predictions per condition")
            if not set(predictions["true_label"]).issubset({"daisy", "dandelion", "roses", "sunflowers", "tulips"}):
                errors.append("predictions.csv contains an unknown true label")

    per_class_path = ROOT / "results" / "final" / "per_class_metrics.csv"
    if per_class_path.exists():
        per_class = pd.read_csv(per_class_path)
        if len(per_class) != 49 * 5:
            errors.append("per_class_metrics.csv must contain 49 × 5 rows")
        elif per_class.duplicated(["condition_id", "class"]).any():
            errors.append("per_class_metrics.csv contains duplicate condition/class pairs")

    stats_path = ROOT / "results" / "final" / "statistical_tests.csv"
    if stats_path.exists():
        stats = pd.read_csv(stats_path)
        required_stats = {
            "condition_id", "baseline_condition_id", "metric", "difference", "ci_lower", "ci_upper",
            "mcnemar_p_value_raw", "mcnemar_p_value_holm", "mcnemar_reject_h0_0_05",
        }
        if len(stats) != 33 * 2:
            errors.append(f"statistical_tests.csv must contain 66 paired rows; found {len(stats)}")
        if not required_stats.issubset(stats.columns):
            errors.append(f"statistical_tests.csv missing columns: {sorted(required_stats.difference(stats.columns))}")
        elif (stats["mcnemar_p_value_holm"] + 1e-15 < stats["mcnemar_p_value_raw"]).any():
            errors.append("Holm-adjusted p-values cannot be smaller than raw p-values")

    confusion_dir = ROOT / "results" / "final" / "confusion_matrices"
    confusion_csvs = sorted(confusion_dir.glob("*.csv")) if confusion_dir.exists() else []
    if len(confusion_csvs) != 49:
        errors.append(f"Expected 49 confusion-matrix CSV files; found {len(confusion_csvs)}")
    else:
        for path in confusion_csvs:
            matrix = pd.read_csv(path, index_col=0)
            if matrix.shape != (5, 5) or int(matrix.to_numpy().sum()) != 550:
                errors.append(f"Invalid confusion matrix: {path.relative_to(ROOT)}")
                break

    metadata_path = ROOT / "models" / "model_metadata.json"
    if metadata_path.exists():
        metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
        required_model_fields = {
            "status", "model_file", "model_sha256", "model_size_bytes", "tensorflow_version",
            "input_shape", "output_shape", "class_names", "train_split_sha256",
            "validation_split_sha256", "seed", "timestamp_utc", "training_duration_seconds",
        }
        if metadata.get("status") != "FULL_RUN_COMPLETE" or not metadata.get("model_sha256"):
            errors.append("model_metadata.json does not prove a FULL_RUN checkpoint")
        missing_model_fields = sorted(required_model_fields.difference(metadata))
        if missing_model_fields:
            errors.append(f"model_metadata.json missing fields: {missing_model_fields}")

    locked_path = ROOT / "configs" / "locked_enhancement_params.json"
    if locked_path.exists():
        locked = read_json(locked_path)
        locked_metadata = locked.get("_metadata", {})
        locked_parameters = locked.get("parameters", {})
        if locked_metadata.get("selection_split") != "validation" or locked_metadata.get("quick_run") is not False:
            errors.append("Locked enhancement parameters do not prove Validation-only FULL_RUN tuning")
        expected_lock = {"metric": "macro_f1", "tie_break_1": "ssim", "tie_break_2": "latency_ms_per_image"}
        for key, value in expected_lock.items():
            if locked_metadata.get(key) != value:
                errors.append(f"Locked enhancement metadata mismatch: {key}")
        for key in ("candidate_grid", "timestamp_utc", "validation_split_sha256", "model_sha256", "seed"):
            if not locked_metadata.get(key):
                errors.append(f"Locked enhancement metadata missing: {key}")
        if len(locked_parameters) != 33:
            errors.append(f"Locked enhancement parameter map must contain 33 entries; found {len(locked_parameters)}")

    notebook_status, notebook_errors = inspect_notebook(require_executed=True, require_full_run=True)
    errors.extend(f"Notebook gate: {item}" for item in notebook_errors)

    submission_metadata_path = ROOT / "configs" / "submission_metadata.json"
    if require_submission and submission_metadata_path.exists():
        admin = read_json(submission_metadata_path)
        required_admin = ("instructor", "group", "members", "submission_date")
        if any(not admin.get(key) for key in required_admin) or len(admin.get("members", [])) != 3:
            errors.append("submission_metadata.json must contain instructor, group, three members and submission_date")
        forbidden_placeholders = ("chưa cung cấp", "placeholder", "tbd", "unknown")
        if any(token in json.dumps(admin, ensure_ascii=False).lower() for token in forbidden_placeholders):
            errors.append("submission_metadata.json still contains placeholder values")

    deployment_path = ROOT / "artifacts" / "deployment_verification.json"
    if require_submission and deployment_path.exists():
        deployment = read_json(deployment_path)
        required_deployment = ("url", "commit", "model_sha256", "verified_at_utc", "screenshot")
        if any(not deployment.get(key) for key in required_deployment):
            errors.append("deployment_verification.json is missing URL/commit/model checksum/time/screenshot evidence")
        elif not (ROOT / deployment["screenshot"]).exists():
            errors.append("Deployment screenshot path does not exist")

    smoke_path = ROOT / "artifacts" / "app_smoke_test.json"
    if require_submission and smoke_path.exists():
        smoke = read_json(smoke_path)
        if smoke.get("status") != "PASS" or smoke.get("model_sha256") != (read_json(metadata_path).get("model_sha256") if metadata_path.exists() else None):
            errors.append("app_smoke_test.json does not prove PASS with the final model checksum")

    manifest_path = ROOT / "results" / "final" / "manifest.json"
    if manifest_path.exists():
        manifest = read_json(manifest_path)
        if manifest.get("condition_count") != 49 or manifest.get("prediction_rows") != 26950:
            errors.append("Final manifest row counts do not match the 49 × 550 contract")
        for relative, expected_hash in manifest.get("files", {}).items():
            artifact = ROOT / relative
            if not artifact.exists() or sha256_file(artifact) != expected_hash:
                errors.append(f"Manifest checksum mismatch: {relative}")
                break

    model_path = ROOT / "models" / "mobilenetv2_flowers.keras"
    if model_path.exists():
        probe = subprocess.run(
            [sys.executable, "-c", "from pathlib import Path; from src.inference import ModelService; ModelService(Path('models/mobilenetv2_flowers.keras')); print('MODEL_LOAD_PASS')"],
            cwd=ROOT,
            check=False,
            capture_output=True,
            text=True,
            timeout=180,
        )
        if probe.returncode != 0 or "MODEL_LOAD_PASS" not in probe.stdout:
            errors.append(f"Fresh-process model load failed: {probe.stderr[-500:]}")

    history_path = ROOT / "artifacts" / "training" / "history.csv"
    if history_path.exists():
        history = pd.read_csv(history_path)
        required_history = {"stage", "epoch", "loss", "accuracy", "val_loss", "val_accuracy", "learning_rate", "duration_seconds"}
        if not required_history.issubset(history.columns):
            errors.append(f"Training history missing columns: {sorted(required_history.difference(history.columns))}")
        elif set(history["stage"]) != {"head", "fine_tune"}:
            errors.append("Training history must contain head and fine_tune stages")

    environment_path = ROOT / "artifacts" / "environment.json"
    if environment_path.exists():
        environment = read_json(environment_path)
        for key in ("python", "platform", "tensorflow", "cpu", "gpu_devices", "seed", "command", "duration_seconds", "created_at_utc"):
            if key not in environment:
                errors.append(f"Environment manifest missing: {key}")

    if require_submission:
        canonical_members = {
            "24100358": "Nguyễn Tùng Dương",
            "24100065": "Trịnh Ngọc Nga",
            "24106898": "Trương Việt Thành",
        }
        visible_paths = [ROOT / "README.md", ROOT / "docs" / "DATA_CARD.md", ROOT / "docs" / "MODEL_CARD.md"]
        visible_paths.extend([ROOT / "docs" / name for name in ("REPORT_FINAL.docx", "SLIDES_FINAL.pptx", "ASSIGNMENT_FINAL.xlsx")])
        texts = []
        for path in visible_paths:
            if not path.exists():
                continue
            texts.append(path.read_text(encoding="utf-8") if path.suffix.lower() == ".md" else office_visible_text(path))
        visible = "\n".join(texts)
        forbidden_admin = re.compile(r"chưa (?:được )?cung cấp|\[\s*(?:điền|họ tên)|\bTBD\b|\bunknown\b|thành viên [123]", re.IGNORECASE)
        if forbidden_admin.search(visible):
            errors.append("Visible final artifacts still contain an administrative placeholder")
        compact = re.sub(r"\s+", "", visible)
        for student_id, name in canonical_members.items():
            if student_id not in visible or re.sub(r"\s+", "", name) not in compact:
                errors.append(f"Visible final artifacts do not contain canonical member {student_id} — {name}")

        try:
            from openpyxl import load_workbook

            book = load_workbook(ROOT / "docs" / "ASSIGNMENT_FINAL.xlsx", data_only=False, read_only=True)
            values = [cell.value for sheet in book.worksheets for row in sheet.iter_rows() for cell in row if cell.value is not None]
            formulas = [value for value in values if isinstance(value, str) and value.startswith("=")]
            if not any("COUNTA(A10:A12)" in value.replace("$", "") for value in formulas):
                errors.append("Assignment workbook does not contain semantic member_count formula over A10:A12")
            if not any("SUM(D10:D12)" in value.replace("$", "") for value in formulas):
                errors.append("Assignment workbook does not contain semantic allocation formula over D10:D12")
        except Exception as exc:  # noqa: BLE001
            errors.append(f"Assignment workbook semantic validation failed: {exc}")

    forbidden = re.compile(r"API_BASE_URL|CORSMiddleware|requests\.post\([^\n]*pipeline|docker compose up", re.IGNORECASE)
    violations: list[str] = []
    allowed_suffixes = {".md", ".py", ".toml", ".yaml", ".yml", ".json", ".txt"}
    for path in ROOT.rglob("*"):
        if not path.is_file() or path.suffix.lower() not in allowed_suffixes:
            continue
        if "legacy" in path.parts or ".venv" in path.parts or path.name in {"validate_project.py", "build_final_text_artifacts.py"}:
            continue
        try:
            if forbidden.search(path.read_text(encoding="utf-8")):
                violations.append(str(path.relative_to(ROOT)))
        except UnicodeDecodeError:
            continue
    if violations:
        errors.append(f"Legacy production strings remain: {violations}")

    return {
        "missing_final_artifacts": missing,
        "legacy_string_violations": violations,
        "notebook": notebook_status,
        "confusion_matrix_csv_count": len(confusion_csvs),
    }, errors


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--require-full-run", action="store_true")
    parser.add_argument("--require-final", action="store_true")
    args = parser.parse_args()
    core_status, errors = core_checks()
    final_status: dict = {}
    if args.require_full_run or args.require_final:
        final_status, final_errors = full_run_checks(require_submission=args.require_final)
        errors.extend(final_errors)
    status = {
        "mode": "require-final" if args.require_final else ("require-full-run" if args.require_full_run else "core"),
        **core_status,
        **final_status,
        "submission_ready": args.require_final and not errors,
        "errors": errors,
    }
    ARTIFACTS.mkdir(parents=True, exist_ok=True)
    (ARTIFACTS / "project_validation.json").write_text(
        json.dumps(status, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(json.dumps(status, ensure_ascii=False, indent=2))
    if errors:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
